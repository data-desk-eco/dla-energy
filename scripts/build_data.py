"""Build the AML Global / DLA Energy contract investigation database.

For each "Fuel Source Location" named in the DLA Energy contract spreadsheet,
pull jet-fuel arrivals (since 2025-01-01) and — for refineries — the upstream
crude. For sites that are storage / transshipment terminals the jet inflows
reveal which refineries fed them, and we pull crude into those refineries too.
"""
from __future__ import annotations

import json
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

import duckdb
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
KPLER = ROOT / "skills/kpler/kpler"
DATA = ROOT / "data"
RAW = DATA / "raw"
DB = DATA / "data.duckdb"
XLSX = DATA / "DLA Energy - AML Contracts.xlsx"

JET_PRODUCT = 1644
CRUDE_GROUP = 1370
SINCE = datetime(2025, 1, 1)
DATA_FLOOR = datetime(2025, 1, 1)  # earliest seaborne data we hold
TODAY = datetime.now()
PAGE = 500
MAX_OFFSET = 9500

# Contract Period of Performance per Kpler installation, from the DLA Energy
# spreadsheet (MM/DD/YYYY in the source, ISO here). Where one installation
# serves two contracts (Petron Bataan: Davao + Zamboanga) the window is the
# union. Crude/jet flows are only counted while a contract was live — clipped
# to the data we hold (from 2025-01-01) and to today.
DLA_CONTRACT_WINDOW = {
    1676:  ("2026-04-01", "2029-03-31"),  # ATT Tanjung Bin → Pago Pago
    1308:  ("2025-08-01", "2029-09-30"),  # Lytton → Canberra/Townsville
    1685:  ("2024-10-21", "2026-09-30"),  # Ocean Point Products → Bridgetown
    2097:  ("2024-10-21", "2026-09-30"),  # Ocean Point → Bridgetown
    7038:  ("2024-02-01", "2027-09-30"),  # NATREF → Gaborone
    6782:  ("2025-08-01", "2029-09-30"),  # Hengyi → Brunei
    1554:  ("2024-02-01", "2027-09-30"),  # Vopak Europoort → Amilcar Cabral
    9617:  ("2025-06-01", "2028-05-31"),  # Mostorod I → Cairo
    9709:  ("2025-06-01", "2028-05-31"),  # Mostorod II → Cairo
    1782:  ("2024-02-01", "2027-09-30"),  # Tema → Kotoka
    2187:  ("2025-08-01", "2029-09-30"),  # Dumai → Soekarno-Hatta
    9595:  ("2025-06-01", "2028-05-31"),  # JOPETROL Zarqa → Marka/Aqaba
    1672:  ("2025-06-01", "2028-05-31"),  # Aqaba Terminal → Marka/Aqaba
    1369:  ("2023-12-01", "2027-09-30"),  # Kipevu → Mombasa/JKIA
    6798:  ("2025-06-01", "2028-05-31"),  # MIDOR → Beirut
    1344:  ("2025-08-01", "2029-09-30"),  # Petronas Melaka → Kuala Lumpur
    1247:  ("2024-02-01", "2027-09-30"),  # Cepsa Huelva → Rabat
    1266:  ("2024-01-01", "2027-09-30"),  # Reliance Jamnagar → Abuja
    2385:  ("2025-06-01", "2028-05-31"),  # OQ MAF Refinery → Muscat
    1355:  ("2025-06-01", "2028-05-31"),  # OQ MAF terminal → Muscat
    4720:  ("2023-08-01", "2029-09-30"),  # Petron Bataan → Davao + Zamboanga
    11317: ("2025-10-01", "2029-03-31"),  # Dangote → San Juan
    1246:  ("2024-02-01", "2027-09-30"),  # SAR M'Bao → Dakar
    1374:  ("2025-06-02", "2030-06-01"),  # IRPC Rayong → Phuket
    4341:  ("2024-02-01", "2027-09-30"),  # STIR Bizerte → Tunis
    1384:  ("2025-08-01", "2029-09-30"),  # Sinopec Hainan → Noi Bai
}


def clip_window(start_iso: str, end_iso: str) -> tuple[datetime, datetime]:
    """Clip a contract window to the data we actually hold."""
    s = max(datetime.fromisoformat(start_iso), DATA_FLOOR)
    e = min(datetime.fromisoformat(end_iso), TODAY)
    return s, e

# Manually mapped from the DLA spreadsheet's (Source, Fuel Source Location)
# columns to Kpler installation IDs. Landlocked sites (Barauni, Fergana,
# Zambia, Sasolburg is pipeline-fed but tracked here as the destination
# refinery) and shut-down sites (Mohammedia/SAMIR for Vivo Maroc) are noted
# but the latter are omitted from the Kpler pulls.
#
# Some DLA rows map to multiple installations (e.g. CORC = Mostorod I + II);
# some installations serve multiple DLA contracts.
DLA_SITES = [
    # (kpler_id, role, source_label, location_label, country_code, refueler, airports)
    (1676,  "terminal", "ATT Tanjung Bin Terminal", "Tanjung Bin", "MY",
        "PACIFIC ISLAND ENERGY", "Pago Pago (American Samoa)"),
    (1308,  "refinery", "Lytton Refinery (Ampol)", "Brisbane", "AU",
        "AMPOL", "Canberra; Townsville"),
    (1685,  "terminal", "Ocean Point Terminals (St Croix Products)", "St Croix", "VI",
        "AML GLOBAL", "Bridgetown (Barbados)"),
    (2097,  "terminal", "Ocean Point Terminals (St Croix)", "St Croix", "VI",
        "AML GLOBAL", "Bridgetown (Barbados)"),
    (7038,  "refinery", "NATREF (Sasol/TotalEnergies)", "Sasolburg", "ZA",
        "PUMA ENERGY", "Gaborone (Botswana)"),
    (6782,  "refinery", "Hengyi Pulau Muara Besar", "Pulau Muara Besar", "BN",
        "GLAMCO UDARA", "Brunei Apt"),
    (1554,  "terminal", "Vopak Terminal Europoort", "Rotterdam", "NL",
        "ENACOL", "Amilcar Cabral (Cape Verde)"),
    (9617,  "refinery", "Mostorod I (CORC)", "Cairo", "EG",
        "MISER PETROLEUM", "Cairo Intl"),
    (9709,  "refinery", "Mostorod II (CORC/ERC)", "Cairo", "EG",
        "MISER PETROLEUM", "Cairo Intl"),
    (1782,  "refinery", "Tema Oil Refinery", "Tema", "GH",
        "PUMA ENERGY", "Kotoka (Accra)"),
    (2187,  "refinery", "Pertamina Dumai Refinery", "Riau", "ID",
        "PERTAMINA AVIATION", "Soekarno-Hatta (Jakarta)"),
    (9595,  "refinery", "JOPETROL Zarqa Refinery", "Zarqa", "JO",
        "JORDAN PETROLEUM", "Marka; Aqaba"),
    (1672,  "terminal", "Aqaba Terminal (JOPETROL crude port)", "Aqaba", "JO",
        "JORDAN PETROLEUM", "Marka; Aqaba"),
    (1369,  "terminal", "Kipevu Oil Storage Facility", "Mombasa", "KE",
        "OLA ENERGY KENYA", "Mombasa Moi; JKIA Nairobi"),
    (6798,  "refinery", "MIDOR Refinery", "Alexandria (Amerya)", "EG",
        "FUEL & AVIATION TRADING", "Beirut Rafic Hariri"),
    (1344,  "refinery", "Petronas Melaka Refinery (MRC)", "Melaka", "MY",
        "PETRONAS DAGANGAN", "Kuala Lumpur"),
    (1247,  "refinery", "Cepsa La Rabida (Huelva)", "Huelva", "ES",
        "OLA ENERGY", "Rabat Sale"),
    (1266,  "refinery", "Reliance Jamnagar", "Jamnagar", "IN",
        "ASHARAMI SYNERGY", "Abuja Nnamdi Azikiwe"),
    (2385,  "refinery", "OQ Mina Al Fahal Refinery", "Muscat", "OM",
        "AL MAHA PETROLEUM", "Muscat Intl"),
    (1355,  "terminal", "OQ Mina Al Fahal terminal", "Muscat", "OM",
        "AL MAHA PETROLEUM", "Muscat Intl"),
    (4720,  "refinery", "Petron Bataan Refinery", "Limay Bataan", "PH",
        "PETRON DCMJR / DCMJR", "Davao; Zamboanga"),
    (11317, "refinery", "Dangote Petroleum Refinery", "Lagos", "NG",
        "PUMA ENERGY AVIATION", "San Juan (Puerto Rico)"),
    (1246,  "refinery", "SAR M'Bao Refinery", "Dakar", "SN",
        "OLA ENERGY SENEGAL", "Leopold Sedar Senghor (Dakar)"),
    (1374,  "refinery", "IRPC Rayong Refinery", "Rayong", "TH",
        "PTTOR", "Phuket"),
    (4341,  "refinery", "STIR Bizerte Refinery", "Bizerte", "TN",
        "OLA ENERGY TUNISIE", "Tunis Carthage"),
    (1384,  "refinery", "Sinopec Hainan Refinery", "Hainan", "CN",
        "SKYPEC", "Noi Bai (Hanoi)"),
]

# DLA rows that have no realistic Kpler footprint, kept for reference in the
# notebook but not pulled.
UNTRACKED = [
    ("Indeni / Puma Energy storage", "Chongwe", "ZM", "PUMA ENERGY",
        "Kenneth Kaunda (Lusaka)", "Zambia is landlocked, pipeline-fed from Tazama"),
    ("Vivo Energy Maroc", "Marrakech", "MA", "ASE MOROCCO",
        "Marrakech Menara", "SAMIR Mohammedia refinery shut in 2015; only storage remains"),
    ("Puma Energy Tanzania storage", "Dar es Salaam", "TZ", "PUMA ENERGY",
        "Julius Nyerere (Dar)", "No operating refinery in Tanzania"),
    ("Barauni Refinery (Indian Oil)", "Barauni Bihar", "IN", "NEPAL OIL",
        "Tribhuvan (Kathmandu)", "Landlocked refinery; jet pipelined to Nepal"),
    ("Fergana Oil Refinery", "Fergana", "UZ", "SANOAT ENERGTIKA",
        "Tashkent Islam Karimov", "Landlocked; no marine crude trail"),
    ("Puma Energy Napa Napa / Port Moresby", "Port Moresby", "PG",
        "PACIFIC ENERGY AVIATION", "Jacksons / Port Moresby",
        "Napa Napa refinery shut 2024; only LPG terminal on Kpler"),
]


def kpler_trades(*, scope_flag: str, scope_id: int, products: int | None,
                 label: str, cache: Path | None = None) -> list[dict]:
    if cache and cache.exists():
        print(f"  {label}: cache hit ({cache.name})")
        return json.loads(cache.read_text())
    out: list[dict] = []
    offset = 0
    while offset <= MAX_OFFSET:
        cmd = [
            str(KPLER), "trades",
            scope_flag, str(scope_id),
            "--size", str(PAGE),
            "--offset", str(offset),
            "--no-forecasted",
        ]
        if products is not None:
            cmd += ["--products", str(products)]
        proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
        if proc.returncode != 0:
            print(f"  {label} @ offset={offset}: {proc.stderr.strip()[:200]}", file=sys.stderr)
            break
        page = [json.loads(line) for line in proc.stdout.splitlines() if line.strip()]
        if not page:
            break
        out.extend(page)
        oldest = min((p.get("start") or "") for p in page)
        print(f"  {label} offset={offset:>5}  +{len(page):>3}  oldest={oldest[:10]}")
        if oldest and oldest < SINCE.isoformat():
            break
        if len(page) < PAGE:
            break
        offset += PAGE
        time.sleep(0.2)
    if cache:
        cache.write_text(json.dumps(out))
    return out


def flatten(trade: dict) -> dict:
    pco = trade.get("portCallOrigin") or {}
    pcd = trade.get("portCallDestination") or {}
    o_inst = pco.get("installation") or {}
    d_inst = pcd.get("installation") or {}
    o_zone = pco.get("zone") or {}
    d_zone = pcd.get("zone") or {}

    fq_list = trade.get("flowQuantities") or []
    fq = fq_list[0] if fq_list else {}
    confirmed = fq.get("confirmedProduct") or {}
    flow = confirmed.get("flowQuantity") or {}
    grade = (fq.get("closestAncestorGrade") or {})
    commodity = (fq.get("closestAncestorCommodity") or {})
    group = (fq.get("closestAncestorGroup") or {})

    osi = ((trade.get("orgSpecificInfo") or {}).get("default") or {})
    seqs = osi.get("bestTradeLinkSequences") or osi.get("tradeLinkSequences") or []
    seller = buyer = None
    if seqs:
        for link in (seqs[0].get("tradeLinks") or []):
            if not seller and link.get("seller"):
                seller = link["seller"].get("name")
            if not buyer and link.get("buyer"):
                buyer = link["buyer"].get("name")

    vessels = trade.get("vessels") or []
    v0 = vessels[0] if vessels else {}

    return {
        "trade_id": trade.get("id"),
        "status": trade.get("status"),
        "start": trade.get("start"),
        "end": trade.get("end"),
        "origin_installation_id": o_inst.get("id"),
        "origin_installation": o_inst.get("name"),
        "origin_port": o_zone.get("name"),
        "origin_country": (o_zone.get("country") or {}).get("name"),
        "dest_installation_id": d_inst.get("id"),
        "dest_installation": d_inst.get("name"),
        "dest_port": d_zone.get("name"),
        "dest_country": (d_zone.get("country") or {}).get("name"),
        "product": commodity.get("name"),
        "grade": grade.get("name"),
        "product_group": group.get("name"),
        "mass_t": flow.get("mass"),
        "volume_bbl": flow.get("volume"),
        "seller": seller,
        "buyer": buyer,
        "vessel_name": v0.get("name"),
        "vessel_imo": v0.get("imo"),
    }


def write_table(con, name: str, rows: list[dict]) -> None:
    con.execute(f"DROP TABLE IF EXISTS {name}")
    if not rows:
        con.execute(f"CREATE TABLE {name} (trade_id BIGINT)")
        return
    raw_path = RAW / f"{name}.json"
    raw_path.write_text(json.dumps(rows))
    con.execute(f"CREATE TABLE {name} AS SELECT * FROM read_json_auto('{raw_path}')")
    con.execute(f"ALTER TABLE {name} ALTER start TYPE TIMESTAMP USING start::TIMESTAMP")
    con.execute(f"ALTER TABLE {name} ALTER \"end\" TYPE TIMESTAMP USING \"end\"::TIMESTAMP")
    con.execute(f"DELETE FROM {name} WHERE start < TIMESTAMP '2025-01-01'")


def load_contracts() -> list[dict]:
    """Parse the DLA Energy contract spreadsheet into a list of dicts."""
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter)]
    out = []
    for row in rows_iter:
        if not any(row):
            continue
        rec = {h: (str(v).strip() if v is not None else None) for h, v in zip(headers, row)}
        out.append(rec)
    return out


def main():
    RAW.mkdir(parents=True, exist_ok=True)
    if DB.exists():
        DB.unlink()
    con = duckdb.connect(str(DB))

    contracts = load_contracts()
    print(f"\nLoaded {len(contracts)} DLA contracts from spreadsheet")
    raw = RAW / "contracts.json"
    raw.write_text(json.dumps(contracts))
    con.execute(f"CREATE TABLE contracts AS SELECT * FROM read_json_auto('{raw}')")

    # Reference table of the mapped sites (for the notebook to join against).
    sites_rows = [
        {"installation_id": iid, "role": role, "source_label": src,
         "location_label": loc, "country_code": cc, "refueler": ref,
         "airports": ap,
         "period_start": DLA_CONTRACT_WINDOW[iid][0],
         "period_end": DLA_CONTRACT_WINDOW[iid][1]}
        for (iid, role, src, loc, cc, ref, ap) in DLA_SITES
    ]
    untracked_rows = [
        {"source_label": s, "location_label": loc, "country_code": cc,
         "refueler": ref, "airports": ap, "note": note}
        for (s, loc, cc, ref, ap, note) in UNTRACKED
    ]
    (RAW / "sites.json").write_text(json.dumps(sites_rows))
    (RAW / "untracked.json").write_text(json.dumps(untracked_rows))
    con.execute(f"CREATE TABLE dla_sites AS SELECT * FROM read_json_auto('{RAW / 'sites.json'}')")
    con.execute(f"CREATE TABLE dla_untracked AS SELECT * FROM read_json_auto('{RAW / 'untracked.json'}')")

    # ---------------- Jet inflows to each DLA installation ----------------
    print(f"\n[1/3] Jet inflows to {len(DLA_SITES)} DLA-listed installations")
    jet_all: list[dict] = []
    seen_jet: set = set()
    for iid, role, src, *_ in DLA_SITES:
        page = kpler_trades(
            scope_flag="--to-installations", scope_id=iid,
            products=JET_PRODUCT, label=f"jet→{src[:40]} ({iid})",
            cache=RAW / f"jet_{iid}.json",
        )
        for t in page:
            tid = t.get("id")
            if tid and tid not in seen_jet:
                seen_jet.add(tid)
                jet_all.append(t)
    jet_flat = [flatten(t) for t in jet_all]
    write_table(con, "jet_trades", jet_flat)
    n_jet = con.execute("SELECT COUNT(*) FROM jet_trades").fetchone()[0]
    print(f"  wrote jet_trades: {n_jet} rows")

    # ---------------- Crude inflows ----------------
    # Direct refineries from DLA list:
    refinery_targets: dict[int, str] = {}
    terminal_ids = set()
    for iid, role, src, *_ in DLA_SITES:
        if role == "refinery":
            refinery_targets[iid] = src
        else:
            terminal_ids.add(iid)

    # Storage terminals: trace one step upstream via jet inflows to identify
    # the refineries that supplied them, and pull crude there too.
    upstream_from_terminals: dict[int, str] = {}
    for row in jet_flat:
        if row.get("dest_installation_id") in terminal_ids:
            up_id = row.get("origin_installation_id")
            up_name = row.get("origin_installation")
            if up_id and up_id not in refinery_targets and up_id not in upstream_from_terminals:
                upstream_from_terminals[up_id] = up_name or f"installation {up_id}"

    crude_targets = {**refinery_targets, **upstream_from_terminals}
    print(f"\n[2/3] Crude inflows to {len(crude_targets)} refineries "
          f"({len(refinery_targets)} from DLA list + "
          f"{len(upstream_from_terminals)} upstream of DLA terminals)")

    crude_all: list[dict] = []
    seen_crude: set = set()
    for iid, name in crude_targets.items():
        page = kpler_trades(
            scope_flag="--to-installations", scope_id=iid,
            products=CRUDE_GROUP, label=f"crude→{name[:40]} ({iid})",
            cache=RAW / f"crude_{iid}.json",
        )
        for t in page:
            tid = t.get("id")
            if tid and tid not in seen_crude:
                seen_crude.add(tid)
                crude_all.append(t)
    crude_flat = [flatten(t) for t in crude_all]
    write_table(con, "crude_trades", crude_flat)
    n_crude = con.execute("SELECT COUNT(*) FROM crude_trades").fetchone()[0]
    print(f"  wrote crude_trades: {n_crude} rows")

    # ---------------- Clip both tables to contract periods ----------------
    # Jet dests are DLA-listed sites — use the site's own contract window.
    jet_win = {iid: clip_window(*DLA_CONTRACT_WINDOW[iid])
               for iid in DLA_CONTRACT_WINDOW}
    term_win = {iid: jet_win[iid] for iid in terminal_ids}

    # Upstream refineries feeding a DLA terminal inherit that terminal's
    # window (union across terminals they feed), counting only jet that
    # actually arrived while the terminal contract was live.
    upstream_win: dict[int, tuple[datetime, datetime]] = {}
    for row in jet_flat:
        tid = row.get("dest_installation_id")
        if tid not in term_win:
            continue
        up = row.get("origin_installation_id")
        if not up:
            continue
        ws, we = term_win[tid]
        end = row.get("end")
        end_dt = datetime.fromisoformat(end[:19]) if end else None
        if end_dt is None or end_dt < ws or end_dt > we:
            continue
        if up in upstream_win:
            s0, e0 = upstream_win[up]
            upstream_win[up] = (min(s0, ws), max(e0, we))
        else:
            upstream_win[up] = (ws, we)

    # Crude dests: a DLA refinery uses its own window; an upstream refinery
    # uses its inherited terminal window. An installation that is both
    # (e.g. Reliance feeds both Abuja directly and Ocean Point) gets the union.
    crude_win: dict[int, tuple[datetime, datetime]] = {}
    for iid in refinery_targets:
        if iid in jet_win:
            crude_win[iid] = jet_win[iid]
    for up, (ws, we) in upstream_win.items():
        if up in crude_win:
            s0, e0 = crude_win[up]
            crude_win[up] = (min(s0, ws), max(e0, we))
        else:
            crude_win[up] = (ws, we)

    def windows_table(name: str, win: dict[int, tuple[datetime, datetime]]):
        rows = [{"inst_id": iid,
                 "win_start": s.isoformat(), "win_end": e.isoformat(),
                 "win_years": max((e - s).days, 0) / 365.25}
                for iid, (s, e) in win.items()]
        (RAW / f"{name}.json").write_text(json.dumps(rows))
        con.execute(f"DROP TABLE IF EXISTS {name}")
        con.execute(f"CREATE TABLE {name} AS SELECT * FROM read_json_auto('{RAW / f'{name}.json'}')")
        con.execute(f"ALTER TABLE {name} ALTER win_start TYPE TIMESTAMP USING win_start::TIMESTAMP")
        con.execute(f"ALTER TABLE {name} ALTER win_end TYPE TIMESTAMP USING win_end::TIMESTAMP")

    windows_table("jet_windows", jet_win)
    windows_table("crude_windows", crude_win)

    con.execute("""
        DELETE FROM jet_trades j WHERE NOT EXISTS (
          SELECT 1 FROM jet_windows w
          WHERE w.inst_id = j.dest_installation_id
            AND j."end" >= w.win_start AND j."end" <= w.win_end)
    """)
    con.execute("""
        DELETE FROM crude_trades c WHERE NOT EXISTS (
          SELECT 1 FROM crude_windows w
          WHERE w.inst_id = c.dest_installation_id
            AND c."end" >= w.win_start AND c."end" <= w.win_end)
    """)
    n_jet2 = con.execute("SELECT COUNT(*) FROM jet_trades").fetchone()[0]
    n_crude2 = con.execute("SELECT COUNT(*) FROM crude_trades").fetchone()[0]
    print(f"  clipped to contract periods: jet {n_jet} → {n_jet2}, "
          f"crude {n_crude} → {n_crude2}")

    # ---------------- Refinery nameplate capacity macro ----------------
    # Annual crude throughput at nameplate × ~85% utilisation, in kt/year.
    # Used to estimate the share of a refinery's intake that Kpler's
    # marine data can see vs the unobserved pipeline/domestic portion.
    # US Gulf refineries are mostly pipeline-fed (Permian/Eagle Ford/
    # Bakken light tight oil) — marine data captures <5% of their slate.
    # Coastal Asian refineries (Reliance, Hengyi, Sinopec Hainan) are
    # mostly marine. Numbers from EIA, IEA, JODI, company filings.
    con.execute("""
        CREATE OR REPLACE MACRO refinery_capacity_kt_yr(name) AS
          CASE name
            WHEN 'CITGO Lake Charles Refinery'   THEN 21100
            WHEN 'ExxonMobil Baton Rouge Refinery' THEN 25900
            WHEN 'Marathon Texas City Refinery'  THEN 31300
            WHEN 'Valero Port Arthur Refinery'   THEN 19600
            WHEN 'ExxonMobil Baytown Refinery'   THEN 29000
            WHEN 'Motiva Port Arthur'            THEN 31900
            WHEN 'Marathon Valero Refineries'    THEN 30000
            WHEN 'Valero Bill E'                 THEN 16000
            WHEN 'Valero Bill W'                 THEN 16000
            WHEN 'Pointe-a-Pierre Refinery'      THEN 8000
            WHEN 'Reficar Refinery'              THEN 8200
            WHEN 'El Palito Refinery'            THEN 7000
            WHEN 'Vertex Mobile'                 THEN 4000
            WHEN 'MOH Corinth Refinery'          THEN 5800
            WHEN 'Thessaloniki Refinery'         THEN 4500
            WHEN 'Repsol Cartagena Refinery'     THEN 11000
            WHEN 'Shell Pernis Refinery'         THEN 20000
            WHEN 'BP Rotterdam'                  THEN 19000
            WHEN 'Shell Europoort'               THEN 21000
            WHEN 'Sasol Augusta'                 THEN 10000
            WHEN 'Ineos Grangemouth Refinery'    THEN 10000
            WHEN 'Jamnagar Refinery'             THEN 69500
            WHEN 'Vadinar Refinery'              THEN 20100
            WHEN 'New Mangalore Refinery'        THEN 15000
            WHEN 'Haldia Terminal'               THEN 7700
            WHEN 'BORL Jamnagar Oil Terminal'    THEN 6200
            WHEN 'Hengyi Refinery'               THEN 6700
            WHEN 'Dangote Refinery'              THEN 32300
            WHEN 'Sinopec Hainan'                THEN 7940
            WHEN 'Petronas Melaka Refinery'      THEN 8440
            WHEN 'Tema Oil Refinery'             THEN 2235
            WHEN 'Rayong IRPC Refinery'          THEN 10700
            WHEN 'MIDOR Refinery'                THEN 4960
            WHEN 'Mostorod Refinery I'           THEN 7050
            WHEN 'Mostorod Refinery II'          THEN 3970
            WHEN 'La Rabida'                     THEN 10900
            WHEN 'Pertamina Dumai'               THEN 8440
            WHEN 'Cilacap Refinery'              THEN 17400
            WHEN 'NATREF Refinery'               THEN 5360
            WHEN 'M''Bao Oil Refinery'           THEN 1490
            WHEN 'Zarqa Refinery'                THEN 4470
            WHEN 'Bizerte'                       THEN 1840
            WHEN 'Lytton Refinery'               THEN 5410
            WHEN 'MAF Refinery'                  THEN 5810
            WHEN 'Petron Bataan Refinery'        THEN 8930
            WHEN 'Sohar Refinery'                THEN 10900
            WHEN 'Ruwais Refinery'               THEN 41700
            WHEN 'Duqm Refinery'                 THEN 11200
            WHEN 'Yanbu Refinery'                THEN 19500
            WHEN 'Jubail Industrial Port'        THEN 19400
            WHEN 'Petro Rabigh'                  THEN 19900
            WHEN 'Sitra Refinery'                THEN 13900
            WHEN 'MAA Refinery'                  THEN 24900
            WHEN 'Marsa El Brega Refinery'       THEN 3500
            WHEN 'El Nasr Refinery'              THEN 6000
            WHEN 'Marifu Refinery'               THEN 6700
            WHEN 'Nippon Mizushima Refinery A'   THEN 13800
            WHEN 'Takaishi Osaka Refinery'       THEN 6100
            WHEN 'KNOC Daesan'                   THEN 28800
            WHEN 'Hyundai Daesan Refinery'       THEN 19400
            WHEN 'S-Oil Onsan'                   THEN 33200
            WHEN 'SK Ulsan'                      THEN 41700
            WHEN 'KPIC Ulsan'                    THEN 13900
            WHEN 'Aster Bukom'                   THEN 23800
            WHEN 'Horizon SGP'                   THEN 7000
            WHEN 'Tankstore'                     THEN 7000
            WHEN 'Vopak Banyan'                  THEN 7000
            WHEN 'Dalian'                        THEN 20100
            WHEN 'Dalian Petrochemical'          THEN 10500
            WHEN 'Sinopec Tianjin'               THEN 12500
            WHEN 'Qingdao Huangdao'              THEN 8400
            WHEN 'Jinzhou Port'                  THEN 6700
            WHEN 'CNPC Qinzhou Refinery'         THEN 10500
            WHEN 'Beilun Suansha'                THEN 27000
            WHEN 'Sinopec Zhanjiang Zhongke Refinery' THEN 10500
            WHEN 'Huizhou Refinery'              THEN 11800
            WHEN 'Petrochina Jieyang Refinery'   THEN 20100
            WHEN 'FREP Plant'                    THEN 12000
            WHEN 'Port Dickson Refinery'         THEN 7400
            WHEN 'Tanjung Bin Refinery'          THEN 9300
            WHEN 'Luanda Refinery'               THEN 2300
            ELSE NULL
          END
    """)

    # ---------------- Per-grade jet yield macro ----------------
    # Atmospheric-distillation kerosene/jet cut as a fraction of crude
    # volume, drawn from published assays (BP, Equinor, ENI, ADNOC, S&P
    # Platts). Used as a relative weight when attributing a refinery's
    # jet output to its crude inputs by origin — so heavy Venezuelan or
    # Canadian crudes (low jet cut) contribute less to the "jet barrel"
    # than light sweet Murban or ESPO.
    #
    # These are simple atmospheric cuts; deep-conversion refineries (e.g.
    # Reliance Jamnagar) lift the absolute jet yield via hydrocracking,
    # but the relative weighting between grades still holds.
    print("\n[3/3] Per-grade jet yield + refinery display-name macros")
    con.execute("""
        CREATE OR REPLACE MACRO jet_yield(grade, country) AS
          CASE
            WHEN grade = 'Urals'                THEN 0.105
            WHEN grade = 'ESPO'                 THEN 0.125
            WHEN grade = 'Sokol'                THEN 0.145
            WHEN grade = 'SBL'                  THEN 0.125
            WHEN grade = 'ARCO'                 THEN 0.090
            WHEN grade = 'Varandey'             THEN 0.105
            WHEN grade = 'KEBCO'                THEN 0.105
            WHEN grade = 'Western Russia Crude' THEN 0.105
            WHEN grade = 'Novy Port'            THEN 0.100
            WHEN grade = 'Kaliningrad'          THEN 0.110
            WHEN grade = 'Sak Bl.'              THEN 0.125
            WHEN grade = 'CPC Russia'           THEN 0.110
            WHEN grade = 'CPC'                  THEN 0.110
            WHEN grade = 'CPC Kazakhstan'       THEN 0.110
            WHEN grade = 'Iran'                 THEN 0.100
            WHEN grade = 'Lavan'                THEN 0.130
            WHEN grade = 'Sirri'                THEN 0.110
            WHEN grade = 'South Pars Co.'       THEN 0.500
            WHEN grade = 'Nile'                 THEN 0.140
            WHEN grade = 'Dar'                  THEN 0.135
            WHEN grade = 'Dar/Nile Crude'       THEN 0.140
            WHEN grade = 'Merey'                THEN 0.065
            WHEN grade = 'Boscan'               THEN 0.040
            WHEN grade = 'Hamaca'               THEN 0.060
            WHEN grade = 'Venezuela Crude'      THEN 0.060
            WHEN grade = 'Arab Lt.'             THEN 0.115
            WHEN grade = 'Arab XLt.'            THEN 0.125
            WHEN grade = 'Arab M'               THEN 0.100
            WHEN grade = 'Arab Hy.'             THEN 0.085
            WHEN grade = 'Arab'                 THEN 0.105
            WHEN grade = 'Khafji'               THEN 0.090
            WHEN grade = 'Shaheen'              THEN 0.100
            WHEN grade = 'Murban'               THEN 0.135
            WHEN grade = 'Zakum'                THEN 0.125
            WHEN grade = 'Das'                  THEN 0.130
            WHEN grade = 'Kuwait'               THEN 0.105
            WHEN grade = 'Basrah Med.'          THEN 0.095
            WHEN grade = 'Basrah Hy.'           THEN 0.080
            WHEN grade = 'Oman'                 THEN 0.105
            WHEN grade = 'Johan Sverdrup'       THEN 0.080
            WHEN grade = 'Midland'              THEN 0.140
            WHEN grade = 'Bonny Lt.'            THEN 0.130
            WHEN grade = 'Tupi'                 THEN 0.110
            WHEN grade = 'Buzios'               THEN 0.105
            WHEN grade = 'Mero'                 THEN 0.110
            WHEN grade = 'Sepia'                THEN 0.110
            WHEN grade = 'Atapu'                THEN 0.105
            WHEN grade = 'Sururu'               THEN 0.105
            WHEN grade = 'Peregrino'            THEN 0.065
            WHEN grade = 'Maya'                 THEN 0.060
            WHEN grade = 'Pacific Dilbit'       THEN 0.045
            WHEN grade = 'Cold Lake blend'      THEN 0.045
            WHEN grade = 'TMX'                  THEN 0.050
            WHEN grade = 'Djeno'                THEN 0.095
            WHEN grade = 'Kimanis'              THEN 0.135
            WHEN grade = 'APAC'                 THEN 0.125
            WHEN grade = 'Champion'             THEN 0.130
            WHEN grade = 'Seria Lt.'            THEN 0.140
            WHEN grade = 'Sah Bl.'              THEN 0.145
            WHEN grade = 'Algeria Co.'          THEN 0.500
            WHEN grade = 'Doba Blend'           THEN 0.090
            WHEN grade = 'Lokele'               THEN 0.100
            WHEN grade = 'Dalia'                THEN 0.095
            WHEN grade = 'Pazflor'              THEN 0.085
            WHEN grade = 'Mostarda'             THEN 0.100
            WHEN grade = 'Nemba'                THEN 0.100
            WHEN grade = 'Palanca'              THEN 0.125
            WHEN grade = 'Kissanje'             THEN 0.110
            WHEN grade = 'Clov'                 THEN 0.110
            WHEN grade = 'Hungo'                THEN 0.100
            WHEN grade = 'Girassol'             THEN 0.115
            WHEN grade = 'Cabinda'              THEN 0.115
            WHEN country = 'Russian Federation' THEN 0.105
            WHEN country = 'Iran'               THEN 0.100
            WHEN country = 'Venezuela'          THEN 0.060
            WHEN country = 'Sudan'              THEN 0.135
            WHEN country = 'Saudi Arabia'       THEN 0.105
            WHEN country = 'United Arab Emirates' THEN 0.130
            WHEN country = 'Iraq'               THEN 0.090
            WHEN country = 'Kuwait'             THEN 0.105
            WHEN country = 'Oman'               THEN 0.105
            WHEN country = 'Qatar'              THEN 0.130
            WHEN country = 'Brazil'             THEN 0.105
            WHEN country = 'United States'      THEN 0.135
            WHEN country = 'Nigeria'            THEN 0.120
            WHEN country = 'Norway'             THEN 0.095
            WHEN country = 'Canada'             THEN 0.050
            WHEN country = 'Mexico'             THEN 0.070
            WHEN country = 'Angola'             THEN 0.105
            WHEN country = 'Algeria'            THEN 0.140
            WHEN country = 'Egypt'              THEN 0.105
            WHEN country = 'Argentina'          THEN 0.120
            WHEN country = 'Indonesia'          THEN 0.110
            WHEN country = 'Malaysia'           THEN 0.125
            WHEN country = 'Brunei'             THEN 0.130
            WHEN country = 'China'              THEN 0.110
            WHEN country = 'Cameroon'           THEN 0.095
            WHEN country = 'Colombia'           THEN 0.080
            WHEN country = 'Ecuador'            THEN 0.080
            WHEN country = 'Gabon'              THEN 0.105
            WHEN country = 'Libya'              THEN 0.135
            WHEN country = 'Equatorial Guinea'  THEN 0.110
            WHEN country = 'Congo'              THEN 0.095
            WHEN country = 'Ghana'              THEN 0.110
            WHEN country = 'Guyana'             THEN 0.110
            WHEN country = 'Mauritania'         THEN 0.105
            ELSE 0.110
          END
    """)
    con.execute("""
        CREATE OR REPLACE MACRO refinery_name(name) AS
          CASE name
            WHEN 'Petronas Melaka Refinery' THEN 'Petronas Melaka (MRC)'
            WHEN 'La Rabida'                THEN 'Cepsa La Rabida (Huelva)'
            WHEN 'MAF Refinery'             THEN 'OQ Mina Al Fahal Refinery'
            WHEN 'Mina Al Fahal'            THEN 'OQ Mina Al Fahal terminal'
            WHEN 'NATREF Refinery'          THEN 'NATREF (Sasolburg)'
            WHEN 'M''Bao Oil Refinery'      THEN 'SAR M''Bao (Dakar)'
            WHEN 'Saint Croix'              THEN 'Ocean Point (St Croix)'
            WHEN 'Saint Croix Products'     THEN 'Ocean Point Products (St Croix)'
            WHEN 'ATB'                      THEN 'ATT Tanjung Bin'
            WHEN 'Lytton Refinery'          THEN 'Lytton (Ampol)'
            WHEN 'Tema Oil Refinery'        THEN 'Tema Oil Refinery'
            WHEN 'Pertamina Dumai'          THEN 'Pertamina Dumai'
            WHEN 'Hengyi Refinery'          THEN 'Hengyi (Brunei)'
            WHEN 'MIDOR Refinery'           THEN 'MIDOR (Alexandria)'
            WHEN 'Mostorod Refinery I'      THEN 'CORC Mostorod I'
            WHEN 'Mostorod Refinery II'     THEN 'ERC Mostorod II'
            WHEN 'Zarqa Refinery'           THEN 'JOPETROL Zarqa'
            WHEN 'Aqaba Terminal'           THEN 'Aqaba Terminal'
            WHEN 'Kipevu'                   THEN 'Kipevu (Mombasa)'
            WHEN 'Vopak Europoort'          THEN 'Vopak Europoort'
            WHEN 'Jamnagar Refinery'        THEN 'Reliance Jamnagar'
            WHEN 'Petron Bataan Refinery'   THEN 'Petron Bataan'
            WHEN 'Dangote Refinery'         THEN 'Dangote (Lagos)'
            WHEN 'Rayong IRPC Refinery'     THEN 'IRPC Rayong'
            WHEN 'Bizerte'                  THEN 'STIR Bizerte'
            WHEN 'Sinopec Hainan'           THEN 'Sinopec Hainan'
            ELSE name
          END
    """)

    print("\nDone.")
    con.close()


if __name__ == "__main__":
    main()
